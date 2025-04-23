import pandas as pd
import itksn
import csv
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from openpyxl import Workbook
import os

def enquiry(list):
    if len(list) == 0:
        return 0
    else:
        return 1
    
def auto_column_widths(data, min_width=40, default_width=100):
    """
    Determine column widths based on whether columns are mostly empty.
    Empty columns will be narrower.
    """
    col_widths = []
    col_count = max(len(row) for row in data)

    for col in range(col_count):
        is_empty = all((len(row) <= col or row[col].strip() == "") for row in data)
        col_widths.append(min_width if is_empty else default_width)

    return col_widths

def csv_to_pdf(csv_file_path, output_pdf_path=None):

    LONG_LANDSCAPE = (24 * inch, 11 * inch)

    if output_pdf_path is None:
        base = os.path.splitext(csv_file_path)[0]
        output_pdf_path = base + ".pdf"

    # Read CSV raw
    with open(csv_file_path, newline='') as f:
        reader = csv.reader(f)
        data = [row for row in reader]
    
     # Normalize row lengths by padding with empty strings
    max_cols = max(len(row) for row in data)
    for i, row in enumerate(data):
        if len(row) < max_cols:
            data[i] = row + [""] * (max_cols - len(row))
    
    # Get smart column widths
    col_widths = auto_column_widths(data)

    # Create PDF
    pdf = SimpleDocTemplate(output_pdf_path, 
        pagesize=LONG_LANDSCAPE,
        leftMargin=20,   # ← small left margin (in points; 72 points = 1 inch)
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    table = Table(data, colWidths=col_widths, hAlign='LEFT')

    # Table styling
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d3d3d3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        #("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    pdf.build([table])
    print(f"PDF saved to: {output_pdf_path}")



def prepend(list, str):
    # Using format()
    str += '{0}'
    list = [str.format(i) for i in list]
    return(list)


def check_sn(serialNumber):
    try:
        if isinstance(serialNumber, list):
            result = [itksn.parse(serial.encode("utf-8")) for serial in serialNumber]
        else:
            result = itksn.parse(serialNumber.encode("utf-8"))
            print(result)

        print("Serial Numbers checked successfully!")
        return result
    except Exception as e:
        print(f"Error: {e}")
        return None

def create_excel(serialNumbers,alternative_serials):
    wb = Workbook()
    ws = wb.active
    for idx, value in enumerate(serialNumbers, start=1):
        ws.cell(row=idx, column=1, value=value)
        ws.cell(row=idx, column=2, value=alternative_serials[idx-1])
    wb.save("serialNumbers.xlsx")

    print("Excel workbook created and serial numbers added.")