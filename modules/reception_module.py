from modules.db_utils import authenticate_user_itkdb, authenticate_user_mongodb
from openpyxl import Workbook
import itkdb
import datetime
import segno
import json
import os
import sys
import shutil
import itksn

def prepend(list, str):
    # Using format()
    str += '{0}'
    list = [str.format(i) for i in list]
    return(list)

## NOT USED
def create_labels(meta_data):
    qr_filename = str(json.loads(meta_data)["atlas_serial"]) + ".png"
    qrcode = segno.make_qr(meta_data)

    qrcode.save("labels/"+qr_filename)

## NOT USED
def print_labels():
    pass

def enquiry(list):
    if len(list) == 0:
        return 0
    else:
        return 1

def check_file_size(args):
    eos = True
    '''
    img_str = ["png","jpg","jpeg","pdf","eps","gif"]
    for arg_key, value in args.items():
        key = arg_key
    for arg in args[key]:
        size = os.path.getsize(arg)
        for st in img_str:
            if st in arg:
                eos = True
        if size > 64000:
            eos = True
    '''
    return eos

def check_sn(serialNumber):
    try:
        if isinstance(serialNumber, list):
            result = [itksn.parse(serial.encode("utf-8")) for serial in serialNumber]
        else:
            result = itksn.parse(serialNumber.encode("utf-8"))
            print(result)

        print("Serial Numbers checked successfully!")
        return result
    except Exception as e:
        print(f"Error: {e}")
        return None

def get_alternative_serial(serial):
    if isinstance(serial, list):
        print("Enter alternative serial numbers...")
        alt_serial = []
        for i in range(len(serial)):
            enter_serial = input("Enter number: ")
            alt_serial.append(enter_serial)
    else:
        alt_serial = input("Enter alternative serial number: ")

    return alt_serial

def create_excel(serialNumbers,alternative_serials):
    wb = Workbook()
    ws = wb.active
    for idx, value in enumerate(serialNumbers, start=1):
        ws.cell(row=idx, column=1, value=value)
        ws.cell(row=idx, column=2, value=alternative_serials[idx-1])
    wb.save("serialNumbers.xlsx")

    print("Excel workbook created and serial numbers added.")

def get_component_type():
    comp_names = ["DATA FLEX", "POWER FLEX", "RING", "Z-RAY FLEX", "TYPE-0 to PP0"]
    print("Select a component type.\n")
    for k, v in enumerate(comp_names):
        print(f"For {v}, press {k}")

    while True:
        try:
            comp_selection = comp_names[int(input("\nInput Selection: "))]
            break
        except (ValueError, IndexError):
            print("Input was not a valid selection. Try again.")
    print(f"Selected {comp_selection}\n")
    return comp_selection


def get_code_and_function(component):
    '''
    Based on component selection, this function returns the XXYY code for the serial number and the function attribute of the component.
    '''

 
    code_prefix = "PI"

    if "DATA" in component:
            code_suffix = "PG"
            function = "D"
    elif "POWER" in component:
            code_suffix = "PP"
            function = "P"
    elif "RING" in component:
            code_suffix = "RF"
            function = "R"
    elif "Z-RAY FLEX" in component:
            code_suffix = "PG"
    elif "TYPE-0 to PP0" in component:
            code_suffix = "PG"
    else:
            code_suffix = None
    xxyy = str(code_prefix) + str(code_suffix)
    return xxyy


def get_production_status(status=''):
    status_list = ["Prototype","Pre-Production", "Production","Dummy"]
    status_num = [0,1,2,9]
    counter = 0
    if status =='':
        for k, v in enumerate(status_list):
            if k == 3:
                k = 9
            print(f"For {v}, press {k}")
        while True:
            try:  
                selection = input("\nInput Selection: ")
                if int(selection) in status_num:
                    status = selection
                else:
                    raise ValueError
                break
            
            except(ValueError,IndexError):
                print("Invalid Input. Try again.")
                counter = counter +1
                if counter %3 ==0:
                    print("")
                    for k, v in enumerate(status_list):
                        if k == 3:
                            k = 9
                        print(f"For {v}, press {k}")


        print(f"Selected {status}\n")
    else: 
        selection = status_num[status_list.index(status)]

    return selection

def get_N2(xxyy,selection_1='',selection_2=''):
    '''
    N2 is a value in the serial number which is dependent on component placement and the number of modules on the component.
    This gives the user a selection for both component placement and # modules.
    Returns N2
    '''
    code = xxyy[2:4]
    placement_options = ["BARREL", "ENDCAP"]
    module_types = ["TRIPLET", "QUAD", "BOTH"]
    if selection_1 == '' and selection_2 == '':
        print("Select Component Placement.")
       
        for k, v in enumerate(placement_options):
            print(f"For {v}, press {k}")
        while True:
            try:
                selection_1 = input("\nInput Selection: ")
                placement = placement_options[int(selection_1)]
                break
            except (ValueError, IndexError):
                print("Invalid Input. Try again.")
        print(f"Selected {placement}\n")

        print("Select Number of Modules")
        for k, v in enumerate(module_types):
            print(f"For {v}, press {k}")
        while True:
            try:
                selection_2 = input("\nInput Selection: ")
                n_modules = module_types[int(selection_2)]
                if int(selection_1) == 0 and int(selection_2) == 2:
                    print("Error: You've selected BARREL and BOTH. Only RINGS are BOTH. Try again.")
                    raise ValueError
                break
            except (ValueError, IndexError):
                print("Invalid Input. Try again.")
        print(f"Selected {n_modules}\n")
    else:
        selection_1 = placement_options.index(selection_1)
        selection_2 = module_types.index(selection_2)

    if int(selection_1) == 0:
        if int(selection_2) == 0:
            N2 = 0
        elif int(selection_2) == 1:
            N2 = 1
    elif int(selection_1) == 1:
        if int(selection_2) == 0:
            while True:
                try:
                    if code == "PG":
                        r05_list = ["R0 DATA FLEX","R0.5 DATA FLEX"]
                        for k, v in enumerate(r05_list):
                            print(f"For {v}, press {k}")
                        selection_3 = input("\ninput selection: ")
                        if int(selection_3) == 0:
                            N2 = 2
                        if int(selection_3) == 1:
                            N2 = 3
                    if code == "PP":
                        N2 = 5
                    if code == "RF":
                        N2 = 3
                    break
                except ValueError:
                    print("Invalid code. Try again.")
        if int(selection_2) == 1:
            N2=4
        if int(selection_2) == 2:
            while True:
                try:
                    if code == "RF":
                        N2 = 5
                    if code == "PG":
                        N2 = 5
                    break
                except ValueError:
                    print("Invalid code. Try again.")

    return N2, module_types[int(selection_2)]

def get_flavor(comp_type):
    '''
    Prompts user to select the flavor of component.
    '''
    print("Select Component Flavor.")
    if comp_type == "L0_BARREL_DATA_FLEX":
        flavor_options = [0]
    if comp_type == "L0_BARREL_POWER_FLEX":
        flavor_options = [0]
    if comp_type == "L1_BARREL_DATA_FLEX":
        flavor_options = [1, 2, 3, 4]
    if comp_type == "L1_BARREL_POWER_FLEX":
        flavor_options = [1, 2]
    if comp_type == "INTERMEDIATE_RING":
        flavor_options = [0]
    if comp_type == "QUAD_RING_R1":
        flavor_options = [0]
    if comp_type == "COUPLED_RING_R01":
        flavor_options = [0]
    if comp_type == "QUAD_MODULE_Z_RAY_FLEX":
        flavor_options = [0]
    if comp_type == "R0_POWER_JUMPER":
        flavor_options = [2]
    if comp_type == "R0_POWER_T":
        flavor_options = [1]
    if comp_type == "R0_DATA_FLEX":
        flavor_options = [1, 2, 3]
    if comp_type == "R05_DATA_FLEX":
        flavor_options = [1, 2]
    if comp_type == "TYPE0_TO_PP0":
        flavor_options = [1, 2]
    for k, v in enumerate(flavor_options):
        print(f"For {v}, press {k}")
    while True:
        try:
            selection = input("\nInput Selection: ")
            flavor = flavor_options[int(selection)]
            break
        except (ValueError, IndexError):
            print("Invalid Input. Try again.")
    print(f"Selected {flavor}\n")

    return flavor


def get_type(xxyy, N2, module):
    code = xxyy[2:4]
    if str(code) == "PG" and N2 == 0:
        comp_type = "L0_BARREL_DATA_FLEX"
    elif str(code)== "PP" and N2 == 0:
        comp_type = "L0_BARREL_POWER_FLEX"
    elif str(code) == "PG" and N2 == 1:
        comp_type = "L1_BARREL_DATA_FLEX"
    elif str(code) == "PP" and N2 == 1:
        comp_type = "L1_BARREL_POWER_FLEX"
    elif str(code) == "RF" and N2 == 3:
        comp_type = "INTERMEDIATE_RING"
    elif str(code) == "RF" and N2 == 4:
        comp_type = "QUAD_RING_R1"
    elif str(code) == "RF" and N2 == 5:
        comp_type = "COUPLED_RING_R01"
    elif str(code) == "PG" and N2 == 4:
        comp_type = "QUAD_MODULE_Z_RAY_FLEX"
    elif str(code) == "PP" and N2 == 5:
        print("Select Component which R0 type.")
        r0_options = ["R0_POWER_T","R0_POWER_JUMPER"]
        for k, v in enumerate(r0_options):
            print(f"For {v}, press {k}")
        while True:
            try:
                selection = input("\nInput Selection: ")
                r0 = r0_options[int(selection)]
                break
            except (ValueError, IndexError):
                print("Invalid Input. Try again.")
        print(f"Selected {r0}\n")
        comp_type = r0
    elif str(code) == "PG" and N2 == 2:
        comp_type = "R0_DATA_FLEX"
    elif str(code) == "PG" and N2 == 3:
        comp_type = "R05_DATA_FLEX"
    elif str(code) == "PG" and N2 == 5:
        comp_type = "TYPE0_TO_PP0"
    else:
        print("Your selection does not exist! Please retry.")
        os.execv(sys.executable, ['python'] + sys.argv)
    return comp_type

def update_test_type(client,mongo_client,meta_data,test_type):
    component = client.get("getComponent", json={"component": meta_data["serialNumber"]})  
    
    if component["currentStage"]["code"] != test_type:
      print("Updating component stage to", test_type)
      set_stage = {
        "component": meta_data["serialNumber"],
        "stage": test_type,
        "rework": False,
        "comment": "updated stage to connectivity on "+str(datetime.datetime),
        "history": True
      }
      client.post("setComponentStage",json=set_stage)
      print("Stage updated!")

    ''' Update stage locally '''  
    db = mongo_client["local"]["itk_testing"]

    comp = db.find_one({"_id": meta_data['serialNumber']})
    if comp is None:
        print("Component doesn't exit locally, cannot update stage!")
        exit()
    if comp['stage'] != test_type:
        up_stage = {'$set':{
           'stage': test_type
            }
        }
        db.update_one({"_id": meta_data['serialNumber']},up_stage)
    
def upload_attachments(client,attch,meta_data,test_type):
   component = client.get("getComponent", json={"component": meta_data["serialNumber"]})  

   ylist = ['y',"yes","Y","YES"]
   nlist = ['n',"no","N","NO"]
   for x in component["tests"]:
      if x['code'] == test_type:
         numInsp = len(x["testRuns"])
         testRun = x["testRuns"]

   # This is in case any argument is a in a different directory
   altered_attch_list =[]
   for arg_key, value in attch.items():
       key = arg_key
   if "/" in attch[key][0]:
      for image in attch[key]:
         g = image.split("/")
         img_name = g[-1]
         shutil.copy2(image, img_name)
         altered_attch_list.append(img_name)
   
   attch_list = []

   if "/" in attch[key][0]:
      for atch in altered_attch_list:
         shutil.copy(atch,itkdb.data)
         attch_list.append(itkdb.data / atch)
         os.remove(atch)

   else:
      for atch in attch[key]:
         shutil.copy(atch,itkdb.data)
         attch_list.append(itkdb.data / atch)

   data_list = []
   for atch, title in zip(attch_list,altered_attch_list):
      data_list.append({
         "testRun": testRun[numInsp-1]["id"],
         "title": title,
         "description": "Attachment for"+test_type,
         "type": "file",
         "url": atch
      })

   attachment_list = []
   for atch in attch_list:
      attachment_list.append({"data": (atch.name, atch.open("rb"), "image/csv")})

   print("You are about to upload",len(attachment_list), "attachments to " +test_type+" test with run number",numInsp,", do you want to continue? (y or n)")
   ans = input("Answer: ")
   if str(ans) in ylist:
      for data, attachment in zip(data_list, attachment_list):
         client.post("createTestRunAttachment",data=data,files=attachment)
      print("Attachment(s) successfully uploaded!")
   else:
      print("Not uploading photos")


def get_comp_info(client,serialNumber):
    comp_filter = {
        "component": serialNumber
    }
    try:
        component = client.get("getComponent",json=comp_filter)
    except:
        print("Component doesn't exist!")
        os.execv(sys.executable, ['python'] + sys.argv)
    test_types = client.get("listTestTypes",json={"project": component["project"]["code"], "componentType":component["componentType"]["code"]})
    test_list = []
    for test_type in test_types:
        test_list.append(test_type["code"])
    meta_data = {
        "serialNumber": component["serialNumber"],
        "project": component["project"]["code"],
        "user": component["user"],
        "institution": component["institution"]["code"],
        "componentType": component["componentType"]["code"],
        "type": component["type"]["code"],
        "testTypes": test_list
    }
    
    return meta_data
    
def get_template(client,meta_data,test_type):
   ind=0
   for test in meta_data["testTypes"]:
      if str(test) == test_type:
         break
      ind += 1
   rec_filter = {
      'project': meta_data['project'],
      'componentType': meta_data['componentType'],
      'code': meta_data['testTypes'][ind]
   }
   test_template = client.get("generateTestTypeDtoSample",json=rec_filter)
   
   return test_template

def enter_serial_numbers(single=False):
   if single == True:
       print("Enter a single serial number. (20UXXYYN1N2N3nnnn)")
       in_serial = input("\nSerial Number: ")
       return in_serial
   
   ylist = ['y',"yes","Y","YES"]
   nlist = ['n',"no","N","NO"]
   
   while True:
        print("Are you entering a batch? (y or n)")
        answer = input("\nAnswer: ")
        try:
            if answer in nlist:
                print("You are entering a single serial number. Please enter it. (20UXXYYN1N2N3nnnn)")
                in_serial = input("\nSerial Number: ")
                return in_serial
            elif answer in ylist:
                print("Please enter the partial serial. (20UXXYYN1N2N3)")
                partial_serial = input("\nPartial serial number: ")
                print("How many are you deleting?")
                quantity = input("\nTotal amount to delete: ")
                print("Starting from which number to delete from? (1 to 9999)")
                start_num = input("\nStarting number: ")
                num_list = []
                for i in range(int(quantity)):
                    if 0 <= i <= 9999:
                        formatted_num = '{:04d}'.format(int(start_num)+i)
                        num_list.append(formatted_num)
                in_serials = prepend(num_list,partial_serial)
                return in_serials
            else:
                raise ValueError
        except ValueError:
            print("Please enter a valid answer. Yes (y) or no (n)")
        
def format_number(latest_serial,existing_serials,register):
    '''
    This ensures the last four digits of the serial number are in fact four digits when represented as a string
    '''
    existing_serial_list = []
    for serial in existing_serials:
        existing_serial_list.append(serial[10:14])

    answers = ["n","y","no","yes","N","Y","NO","YES"]
    nlist = ["n","no","N","NO"]
    ylist = ['y',"yes","Y","YES"]
    print("Are you entering a batch? (y or n)")
    while True:
        try:
            answer = input("\nAnswer: ")
            if answer not in answers:
                raise ValueError
            break
        except ValueError:
            print("Invalid answer. Only a y or n")
    if answer in nlist:
        while True:
            try:
                print("The latest number is:",latest_serial)
                if register == True:
                    print("The recommended number to enter is:", int(latest_serial)+1)
                while True:
                    print("Enter a number for the component (1 to 9999)")
                    number = input("\nInput Selection: ")
                    num = int(number)
                    if register == True:
                        try:
                            if '{:04d}'.format(num) in existing_serial_list:
                                raise ValueError
                            elif num == 0:
                                raise ValueError
                            else:
                                break
                        except ValueError:
                            print("The number you entered already exists! Please enter a non-existing number.")
                    elif register == False:
                        try:
                            if '{:04d}'.format(num) not in existing_serial_list:
                                raise ValueError
                            else:
                                break
                        except ValueError:
                            print("The number you entered doesn't exist! Please enter an existing number.")

                if 0 <= num <= 9999:
                    formatted_num = '{:04d}'.format(num)
                    return formatted_num
                else:
                    raise ValueError
            except ValueError:
                print("Invalid input. Please provide a valid number.")
    
    if answer in ylist:
        while True:
            try:
                if register == True:
                    print("Enter how many components to register in a batch (2 to 9999)")
                if register == False:
                    print("Enter how many components to delete in a batch (2 to 9999)")
                number = input("\nTotal Amount: ")
                if 2<= int(number) <= 9999:
                    break
            except ValueError:
                print("Invalid input. Please provide a valid number.")
        num = int(number)
        while True:
            print("The latest serial number is:",latest_serial)
            print("Do you wish to start from the latest serial number position? (y or n)")
            ans = input("\nAnswer: ")

            try:
                if ans in ylist:
                    if register == True:
                        start_num = latest_serial+1
                    if register == False: 
                        start_num = latest_serial
                    break
                elif ans in nlist:
                    print("Starting from a different position, please enter starting position")
                    while True:    
                        start_number = input("\nStart Number: ")
                        start_num = int(start_number)
                        if register == True:
                            try:
                                for p in range(num):
                                    if '{:04d}'.format(start_num+p) in existing_serial_list:
                                        raise ValueError
                                break
                            except ValueError:
                                print("The numbers you've entered already exists! Please enter a non-existing serial numbers.")
                        elif register == False:
                            try:
                                for p in range(num):
                                    if '{:04d}'.format(start_num-p) not in existing_serial_list:
                                        raise ValueError
                                break
                            except ValueError:
                                print("The numbers you've entered don't exist! Please enter a existing serial numbers.")
                    break
                else:
                    raise ValueError
            except ValueError:
                print("Invalid input. Try yes (y) or no (n)")

        comp_list = []
        for i in range(num):
            if 0 <= i <= 9999:
                if register == True:
                    formatted_num = '{:04d}'.format(start_num+i)
                    comp_list.append(formatted_num)
                if register == False:
                    formatted_num = '{:04d}'.format(start_num-i)
                    comp_list.append(formatted_num)
            else:
                return "Number out of range. Please provide a number between 1 and 9999."
        return comp_list

def get_existing_serials(client,partial_serial,xxyy,N2,flavor):
    print("Retrieving existing components...")
    project = xxyy[0]
    subproject1 = xxyy[0:2]
    subproject2 = xxyy[2:4]
    comp_type = get_type(xxyy,N2)
    print("The component type you're searching is:", comp_type)
    print("Searching production database for this type...")
    search_filter = {
        "filterMap": {
            "project": project,
            "subproject": [subproject1,subproject2],
            "type": comp_type,
            "institute": "OSU"
        }
    }
    existing_components = client.get("listComponents", json=search_filter)
    if isinstance(existing_components,list):
        print("Total components of type", comp_type,"found is:",len(existing_components))
    else:
        print("Total components of type", comp_type,"found is:",existing_components.total)

    production_status = partial_serial[7] 
    status_list = ["Prototype","Pre-Production", "Production","Dummy"]
    prod_status = 3 if int(production_status) == 9 else int(production_status)
    status = status_list[prod_status]

    existing_osu_components = []
    existing_components_status_flavor = []
    for i in existing_components:
        '''For deleted components'''
        if i['state'] == "deleted":
            continue
        code = str(i["institution"]["code"])
        if code == str("OSU"):
            existing_osu_components.append(i)
            if i["serialNumber"][9] == str(flavor) and i['serialNumber'][7]:
                 existing_components_status_flavor.append(i)
    print("Total components of type", comp_type," of status",status," with flavor", flavor, "is",len(existing_components_status_flavor))

    existing_serials = []
    for component in existing_components_status_flavor:
        if partial_serial in component["serialNumber"]:
            existing_serials.append(component["serialNumber"])
        elif component["serialNumber"] is None:
            pass
    return existing_serials
    
def get_latest_serial(client,xxyy, production_status, N2, flavor, register,comp_type): 
    '''
    Checks data base for existing components with the same first 10 digits.
    Finds largest existing serial number and increments it by 1.
    Returns the new serial number.
    '''
    partial_serial = "20U" + str(xxyy) + str(production_status) + str(N2) + str(flavor)
    print("The partial serial number entered:",partial_serial)
    project = xxyy[0]
    subproject1 = xxyy[0:2]
    subproject2 = xxyy[2:4]

    status_list = ["Prototype","Pre-Production", "Production","Dummy"]
    prod_status = 3 if int(production_status) == 9 else int(production_status)
    status = status_list[prod_status]

    #comp_type = get_type(xxyy,N2)
    print("The component type you're entering is:", comp_type)
    print("Searching production database for this type...")
    search_filter = {
        "filterMap":{
            "project": project,
            "subproject": [subproject1,subproject2],
            "type": comp_type,
            "institute": "OSU"
        }
    }
    existing_components = client.get("listComponents", json=search_filter)

    if isinstance(existing_components,list):
        print("Total components of type", comp_type,"found is:",len(existing_components))
    else:
        print("Total components of type", comp_type,"found is:",existing_components.total)

    existing_osu_components = []
    existing_components_status_flavor = []
    for i in existing_components:
        
        ''' For deleted components '''
        if i['state'] == 'deleted':
            continue

        code = str(i["institution"]["code"])
        if code == str("OSU"):
            existing_osu_components.append(i)
            if i["serialNumber"][9] == str(flavor) and i["serialNumber"][7] == str(production_status):
                 existing_components_status_flavor.append(i)
    print("Total components of type", comp_type,"of status",status,"with flavor", flavor, "is",len(existing_components_status_flavor))

    existing_serials = []
    for component in existing_components_status_flavor:
        if partial_serial in component["serialNumber"]:
            existing_serials.append(component["serialNumber"])
        elif component["serialNumber"] is None:
            pass

    latest_serial = 0
    for serial in existing_serials:
        if len(serial) != 14:
            pass
        else:
            if latest_serial < int(serial[10:14]):
                latest_serial = int(serial[10:14])
            else:
                pass
    new_serial = format_number(latest_serial,existing_serials,register)
    
    if isinstance(new_serial,str):
        if register == True:
            print("New serial number:",partial_serial + new_serial )
        if register == False:
            print("Serial numbers to delete:", partial_serial + new_serial)
        return partial_serial + new_serial
    else:
        serial_list = prepend(new_serial,partial_serial)
        if register == True:
            print("New serial numbers: ", serial_list)
        if register == False:
            print("Serial numbers to delete: ", serial_list)
        return serial_list